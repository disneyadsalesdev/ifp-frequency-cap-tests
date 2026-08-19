"""Add summary sheet and sort August lines by deal size."""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SOURCE = Path(r"c:\Users\syeda012\Downloads\Aug_Lines_8.4.26.xlsx")
OUTPUT = SOURCE

HEADER_ROW = 3  # 1-based openpyxl row with column headers
DATA_START = 4
TITLE = "MLB_Hulu_8.3"
SNAPSHOT = "As of 8/4/2026"

LINE_COLUMNS = [
    "Deal Rank",
    "Recommended Action",
    "Sales Order ID",
    "Sales Order Name",
    "Sales Child LI ID",
    "Sales Child LI Name",
    "Sales Child LI Start Date",
    "Sales Child LI End Date",
    "MC Line Frontload %",
    "MC Line Completions Remaining",
    "MC Line Revenue At Risk",
    "Sales Child LI Net Contracted Cost",
    "Deal Total $",
    "Completions",
    "Pace vs Need %",
]

ORIGINAL_COLUMNS = [
    "Sales Order ID",
    "Sales Order Name",
    "Sales Child LI ID",
    "Sales Child LI Name",
    "Sales Child LI Start Date",
    "Sales Child LI End Date",
    "MC Line Frontload %",
    "MC Line Completions Remaining",
    "MC Line Revenue At Risk",
    "Sales Child LI Net Contracted Cost",
    "Completions",
]


def parse_pct(value):
    if pd.isna(value):
        return None
    try:
        return float(str(value).strip().replace("%", ""))
    except ValueError:
        return None


def load_lines() -> pd.DataFrame:
    xl = pd.ExcelFile(SOURCE)
    if "MLB_Hulu_8.3" in xl.sheet_names:
        df = pd.read_excel(SOURCE, sheet_name="MLB_Hulu_8.3", header=2)
    elif "Lines by Deal Size" in xl.sheet_names:
        raw = pd.read_excel(SOURCE, sheet_name="Lines by Deal Size", header=None)
        if raw.shape[1] >= 30 and pd.notna(raw.iloc[4, 17]):
            # Recover from a prior misaligned write (data starts at column P).
            recovered = raw.iloc[4:, 15:30].copy()
            recovered.columns = LINE_COLUMNS
            df = recovered
        else:
            df = pd.read_excel(SOURCE, sheet_name="Lines by Deal Size", header=3)
    else:
        df = pd.read_excel(SOURCE, sheet_name=0, header=2)

    df = df.dropna(subset=["Sales Order ID"]).copy()
    df = df[pd.to_numeric(df["Sales Order ID"], errors="coerce").notna()].copy()
    df["Sales Order ID"] = df["Sales Order ID"].astype(int)
    df["Sales Child LI ID"] = df["Sales Child LI ID"].astype(int)

    if "Recommended Action" in df.columns:
        df["recommended_action"] = df["Recommended Action"]
        df["priority_rank"] = df["Deal Rank"]
        df["order_cost"] = df["Deal Total $"]
        df["pace_pct"] = df["Pace vs Need %"]
    return df


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    if "recommended_action" in df.columns and "Pace vs Need %" in df.columns:
        return df

    today = pd.Timestamp("2026-08-04")
    df["frontload_pct"] = df["MC Line Frontload %"].apply(parse_pct)
    df["end"] = pd.to_datetime(df["Sales Child LI End Date"])
    df["days_remaining"] = ((df["end"] - today).dt.days + 1).clip(lower=1)
    df["required_daily"] = df["MC Line Completions Remaining"] / df["days_remaining"]
    df["pace_pct"] = (
        df["Completions"] / df["required_daily"].replace(0, pd.NA)
    ) * 100
    df["aug_end"] = df["end"].dt.month == 8

    order_totals = df.groupby("Sales Order ID", as_index=False).agg(
        order_cost=("Sales Child LI Net Contracted Cost", "sum"),
        order_remaining=("MC Line Completions Remaining", "sum"),
        order_risk=("MC Line Revenue At Risk", "sum"),
        max_frontload=("frontload_pct", "max"),
        min_frontload=("frontload_pct", "min"),
        avg_pace=("pace_pct", "mean"),
        order_end=("end", "max"),
        sales_order_name=("Sales Order Name", "first"),
    )

    def order_action(row):
        if row["max_frontload"] >= 400 and row["avg_pace"] < 50:
            return "Custom Pacing"
        if row["min_frontload"] < 400 and row["order_end"].month == 8:
            if (
                row["order_cost"] >= 2500
                or row["order_remaining"] >= 50000
                or row["order_risk"] > 0
                or "CATDAA" in str(row["sales_order_name"])
            ):
                return "400% Frontload"
        return "Monitor"

    order_totals["recommended_action"] = order_totals.apply(order_action, axis=1)
    order_totals["priority_rank"] = (
        order_totals["order_cost"].rank(method="dense", ascending=False).astype(int)
    )

    df = df.merge(
        order_totals[
            [
                "Sales Order ID",
                "order_cost",
                "order_remaining",
                "recommended_action",
                "priority_rank",
            ]
        ],
        on="Sales Order ID",
        how="left",
    )

    df = df.sort_values(
        ["priority_rank", "order_cost", "Sales Child LI Net Contracted Cost"],
        ascending=[True, False, False],
    ).reset_index(drop=True)

    df["Pace vs Need %"] = df["pace_pct"].round(1)
    df["Recommended Action"] = df["recommended_action"]
    df["Deal Rank"] = df["priority_rank"]
    df["Deal Total $"] = df["order_cost"].round(2)
    return df


def build_summary_rows(df: pd.DataFrame) -> list[list]:
    rows: list[list] = []
    rows.append(["August Big Deals - Frontload and Pacing Summary", "", "", "", "", ""])
    rows.append([SNAPSHOT, "", "", "", "", ""])
    rows.append(["", "", "", "", "", ""])
    rows.append(
        [
            "Rule",
            "400% Frontload = still on 10-30% and behind pace",
            "",
            "",
            "",
            "",
        ]
    )
    rows.append(
        [
            "",
            "Custom Pacing = already at 400% frontload but still underwater",
            "",
            "",
            "",
            "",
        ]
    )
    rows.append(["", "", "", "", "", ""])

    action_order = ["Custom Pacing", "400% Frontload", "Monitor"]
    for action in action_order:
        subset = (
            df.groupby(
                ["Sales Order ID", "Sales Order Name", "recommended_action"],
                as_index=False,
            )
            .agg(
                order_cost=("Sales Child LI Net Contracted Cost", "sum"),
                order_remaining=("MC Line Completions Remaining", "sum"),
                line_count=("Sales Child LI ID", "count"),
                order_end=("Sales Child LI End Date", "max"),
                avg_pace=("pace_pct", "mean"),
                frontload=("MC Line Frontload %", "first"),
            )
            .query("recommended_action == @action")
            .sort_values("order_cost", ascending=False)
        )
        if subset.empty:
            continue

        rows.append([action.upper(), "", "", "", "", ""])
        rows.append(
            [
                "Deal Rank",
                "Sales Order ID",
                "Sales Order Name",
                "Deal $",
                "Remaining",
                "End Date",
                "Lines",
                "Frontload",
                "Avg Pace %",
            ]
        )
        for _, r in subset.iterrows():
            end_val = r["order_end"]
            end_str = end_val.strftime("%Y-%m-%d") if hasattr(end_val, "strftime") else str(end_val)
            rows.append(
                [
                    int(df.loc[df["Sales Order ID"] == r["Sales Order ID"], "priority_rank"].iloc[0]),
                    int(r["Sales Order ID"]),
                    r["Sales Order Name"],
                    round(float(r["order_cost"]), 2),
                    int(r["order_remaining"]),
                    end_str,
                    int(r["line_count"]),
                    r["frontload"],
                    round(float(r["avg_pace"]), 1) if pd.notna(r["avg_pace"]) else "",
                ]
            )
        rows.append(["", "", "", "", "", ""])

    custom = df[df["recommended_action"] == "Custom Pacing"]
    frontload = df[df["recommended_action"] == "400% Frontload"]
    rows.append(["TOTALS", "", "", "", "", ""])
    rows.append(
        [
            "Custom Pacing deals",
            len(custom.groupby("Sales Order ID")),
            f"${custom.groupby('Sales Order ID')['Sales Child LI Net Contracted Cost'].sum().sum():,.0f}",
            "",
            "",
            "",
        ]
    )
    rows.append(
        [
            "400% Frontload deals",
            len(frontload.groupby("Sales Order ID")),
            f"${frontload.groupby('Sales Order ID')['Sales Child LI Net Contracted Cost'].sum().sum():,.0f}",
            "",
            "",
            "",
        ]
    )
    return rows


def write_workbook(df: pd.DataFrame) -> None:
    lines_out = df[LINE_COLUMNS].copy()

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        summary_df = pd.DataFrame(build_summary_rows(df))
        summary_df.to_excel(writer, sheet_name="Summary", index=False, header=False)

        sheet_name = "Lines by Deal Size"
        pd.DataFrame([[TITLE]]).to_excel(
            writer, sheet_name=sheet_name, index=False, header=False, startrow=0
        )
        lines_out.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=3
        )

    wb = load_workbook(OUTPUT)
    style_summary(wb["Summary"])
    style_lines(wb["Lines by Deal Size"], len(LINE_COLUMNS))
    wb.save(OUTPUT)


def style_summary(ws) -> None:
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12

    title_fill = PatternFill("solid", fgColor="1F4E79")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    section_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="E2EFDA")
    header_font = Font(bold=True)

    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.fill = title_fill
        cell.font = title_font

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        first = row[0].value
        if isinstance(first, str) and first in {"CUSTOM PACING", "400% FRONTLOAD", "MONITOR", "TOTALS"}:
            for cell in row:
                cell.fill = section_fill
                cell.font = section_font
        if first == "Deal Rank":
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font


def style_lines(ws, col_count: int) -> None:
    widths = [10, 18, 14, 42, 14, 48, 16, 16, 16, 18, 16, 16, 12, 12, 12]
    for idx, width in enumerate(widths[:col_count], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    header_row = 4
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    action_fill = {
        "400% Frontload": PatternFill("solid", fgColor="FCE4D6"),
        "Custom Pacing": PatternFill("solid", fgColor="E2EFDA"),
        "Monitor": PatternFill("solid", fgColor="F2F2F2"),
    }
    for row in range(header_row + 1, ws.max_row + 1):
        action = ws.cell(row=row, column=2).value
        fill = action_fill.get(action)
        if fill:
            ws.cell(row=row, column=2).fill = fill


def main() -> None:
    df = enrich(load_lines())
    write_workbook(df)
    print(f"Updated: {OUTPUT}")
    print(f"Lines: {len(df)}")
    print(
        "Actions:",
        df.groupby("recommended_action")["Sales Order ID"].nunique().to_dict(),
    )


if __name__ == "__main__":
    main()
